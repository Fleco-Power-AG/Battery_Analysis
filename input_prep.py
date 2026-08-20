"""
fill_missing_inputs.py
=======================

Ergaenzt die fehlenden Zeitreihen im Batterie/PV-Analyse-Excel (input.xlsx)
und speichert das Ergebnis als input_lp.xlsx (fertig fuer den oemof.solph-Aufbau).

Was das Skript macht (siehe Funktionen unten):
  1. SwissIX Spotpreis (Zeitreihen!B) via ENTSO-E Transparency Platform API laden,
     UTC -> CET/CEST konvertieren, EUR/MWh -> CHF/kWh konvertieren (via EUR/CHF-Kurs).
  2. Rueckliefertarif / Bezugstarif (Zeitreihen!C/D) je nach Parameter!C5 (Tarifschema)
     berechnen: HT_NT (Wochentag/Uhrzeit-Schema aus Parameter!C6:C12),
     SPOT (SwissIX +/- Auf-/Abschlag aus Parameter!C13/C14),
     oder Zeitreihen (vorhandene historische Werte unveraendert uebernehmen).
  3. PV-Produktionsprofil (Zeitreihen!E, "CH_PV_normiert_1kWp") optional via PVGIS
     laden -- braucht Standort/Ausrichtung, siehe fetch_pv_reference_profile().
  4. SRL-Preise (Zeitreihen!G/H) sind im Original-Excel bereits per Array-Formel
     aus Hilfen_Listen berechnet -- werden hier als fixe Werte "gebacken", damit
     die Endversion ohne Hilfsblatt-Referenzen auskommt.
  5. Speichert alles unter input_lp.xlsx (Parameter/Hilfen_Listen unveraendert,
     Zeitreihen ergaenzt).

WICHTIGER HINWEIS ZUR NETZWERK-ABHAENGIGKEIT:
  Schritt 1 (ENTSO-E) und Schritt 3 (PVGIS) brauchen echten Internetzugang.
  Dieses Skript wurde in einer Sandbox entwickelt, die nur auf Paket-Registries
  (PyPI etc.) zugreifen kann, nicht auf beliebige APIs. Die ENTSO-E-Anbindung ist
  daher gegen die dokumentierte API geschrieben, aber NICHT gegen eine echte
  Live-Antwort getestet worden. Bitte einen kurzen Testlauf mit einem kleinen
  Zeitraum (z.B. 2 Tage) machen, bevor das ganze Jahr geladen wird.

Benoetigte Pakete: pandas, openpyxl, requests
    pip install pandas openpyxl requests
"""

from __future__ import annotations

import datetime as dt
import os
import re
import warnings
from xml.etree import ElementTree as ET

import pandas as pd
import requests
from openpyxl import load_workbook

# --------------------------------------------------------------------------
# Konfiguration
# --------------------------------------------------------------------------

INPUT_PATH = "Inputs.xlsx"
OUTPUT_PATH = "Input_lp.xlsx"

ENTSOE_API_KEY = os.environ.get(
    "ENTSOE_API_KEY", "73d39d94-710e-4a11-9160-9f8d98b6047e"
)
ENTSOE_BASE_URL = "https://web-api.tp.entsoe.eu/api"
# EIC-Code fuer die Schweizer Regelzone (Swissgrid)
ENTSOE_CH_DOMAIN = "10YCH-SWISSGRIDZ"

FX_API_URL = "https://api.frankfurter.app"  # kostenlos, kein Key noetig

# Das Excel zaehlt durchgehend 365*96 = 35040 Viertelstunden pro Jahr, ohne
# Sommerzeit-Sprung/-Dopplung zu beruecksichtigen (Kommentar im Sheet: "in CET").
# Wir behandeln "CET" hier deshalb bewusst als FESTEN Offset UTC+1 (nicht die
# echte, DST-wechselnde Europe/Zurich-Zeitzone) -- das entspricht exakt der
# Zaehlweise im Original-Excel und vermeidet Ambiguous-/Nonexistent-Zeit-Fehler
# beim Sommerzeitwechsel.
CET_FIXED = dt.timezone(dt.timedelta(hours=1))

# Spaltenlayout im Sheet "Zeitreihen" (1-indexiert wie in Excel)
COL_ZEIT = 1
COL_SWISSIX = 2
COL_RUECKLIEFER = 3
COL_BEZUG = 4
COL_PV_NORMIERT = 5
COL_PV_PROFIL = 6
COL_SRL_NEG = 7
COL_SRL_POS = 8


# --------------------------------------------------------------------------
# 1. SwissIX / ENTSO-E
# --------------------------------------------------------------------------

def _strip_ns(tag: str) -> str:
    """Entfernt das XML-Namespace-Praefix von einem Tag-Namen."""
    return tag.split("}", 1)[-1] if "}" in tag else tag


def _parse_entsoe_day_ahead_xml(xml_bytes: bytes) -> pd.Series:
    """
    Parst ein ENTSO-E "Publication_MarketDocument" (Day-ahead prices, A44)
    und gibt eine pandas Series (UTC-indexiert, EUR/MWh) zurueck.
    """
    root = ET.fromstring(xml_bytes)

    # Fehlerantworten haben ein "Acknowledgement_MarketDocument" mit Reason-Text
    root_tag = _strip_ns(root.tag)
    if root_tag == "Acknowledgement_MarketDocument":
        reason = root.find(".//{*}Reason/{*}text")
        msg = reason.text if reason is not None else "unbekannter Fehler"
        raise RuntimeError(f"ENTSO-E API Fehlerantwort: {msg}")

    values = {}
    for ts in root.iter():
        if _strip_ns(ts.tag) != "TimeSeries":
            continue
        for period in ts:
            if _strip_ns(period.tag) != "Period":
                continue
            time_interval = period.find("{*}timeInterval")
            start_str = time_interval.find("{*}start").text
            resolution_str = period.find("{*}resolution").text

            start = pd.Timestamp(start_str, tz="UTC")
            resolution = pd.Timedelta(resolution_str.replace("PT", "").lower())
            # ISO 8601 Dauer wie "PT60M"/"PT15M" -> pandas Timedelta braucht
            # z.B. "60min"/"15min"; obiges .replace ist ein einfacher Spezialfall.
            match = re.match(r"PT(\d+)M", resolution_str)
            if match:
                resolution = pd.Timedelta(minutes=int(match.group(1)))

            for point in period.iter():
                if _strip_ns(point.tag) != "Point":
                    continue
                pos = int(point.find("{*}position").text)
                price = float(point.find("{*}price.amount").text)
                ts_utc = start + (pos - 1) * resolution
                values[ts_utc] = price

    if not values:
        raise RuntimeError(
            "Keine Preis-Punkte in der ENTSO-E-Antwort gefunden -- "
            "moeglich, dass fuer die Schweizer Zone (10YCH-SWISSGRIDZ) "
            "keine Day-Ahead-Preise auf ENTSO-E publiziert werden. "
            "Bitte pruefen (SwissIX wird von EPEX Spot separat berechnet "
            "und ist evtl. nicht 1:1 identisch mit den ENTSO-E-Daten)."
        )

    series = pd.Series(values).sort_index()
    series.index.name = "time_utc"
    return series


def fetch_entsoe_day_ahead_prices(
    start: pd.Timestamp, end: pd.Timestamp, api_key: str = ENTSOE_API_KEY
) -> pd.Series:
    """
    Laedt Day-Ahead-Preise (EUR/MWh) fuer die Schweiz von ENTSO-E, chunked
    nach Monat (ENTSO-E limitiert die Anfragegroesse). Rueckgabe: Series,
    UTC-indexiert.
    """
    chunks = []
    cur = start
    while cur < end:
        chunk_end = min(cur + pd.DateOffset(months=1), end)
        params = {
            "documentType": "A44",
            "in_Domain": ENTSOE_CH_DOMAIN,
            "out_Domain": ENTSOE_CH_DOMAIN,
            "periodStart": cur.strftime("%Y%m%d%H%M"),
            "periodEnd": chunk_end.strftime("%Y%m%d%H%M"),
            "securityToken": api_key,
        }
        resp = requests.get(ENTSOE_BASE_URL, params=params, timeout=60)
        resp.raise_for_status()
        chunks.append(_parse_entsoe_day_ahead_xml(resp.content))
        cur = chunk_end

    return pd.concat(chunks).sort_index()


def fetch_eur_chf_fx_rates(start_date: str, end_date: str) -> pd.Series:
    """Taegliche EUR/CHF-Kurse (Frankfurter/ECB-Referenz), noetig fuer die
    Umrechnung von EUR/MWh (ENTSO-E) nach CHF/kWh."""
    url = f"{FX_API_URL}/{start_date}..{end_date}"
    resp = requests.get(url, params={"from": "EUR", "to": "CHF"}, timeout=30)
    resp.raise_for_status()
    data = resp.json()["rates"]
    fx = pd.Series({pd.Timestamp(d): v["CHF"] for d, v in data.items()}).sort_index()
    return fx


def build_swissix_chf_per_kwh(
    target_index: pd.DatetimeIndex, api_key: str = ENTSOE_API_KEY
) -> pd.Series:
    """
    Baut die SwissIX-Preisreihe in CHF/kWh auf dem Zielindex (Europe/Zurich,
    15-Min-Aufloesung) auf: ENTSO-E (EUR/MWh, UTC) laden, auf Zielindex
    ausrichten (ffill fuer stuendliche Ausgangsdaten), dann mit dem
    Tages-EUR/CHF-Kurs in CHF/kWh umrechnen.
    """
    start_utc = target_index[0].tz_convert("UTC")
    end_utc = target_index[-1].tz_convert("UTC") + pd.Timedelta(minutes=15)

    prices_eur_mwh = fetch_entsoe_day_ahead_prices(start_utc, end_utc, api_key)
    prices_eur_mwh.index = prices_eur_mwh.index.tz_convert(CET_FIXED)

    # Auf den Zielindex ausrichten (falls ENTSO-E nur stuendlich liefert,
    # werden die 4 Viertelstunden je Stunde mit demselben Preis befuellt)
    aligned = prices_eur_mwh.reindex(target_index, method="ffill")

    fx = fetch_eur_chf_fx_rates(
        target_index[0].strftime("%Y-%m-%d"), target_index[-1].strftime("%Y-%m-%d")
    )
    fx_daily = fx.reindex(target_index.normalize().tz_localize(None), method="ffill")
    fx_daily.index = target_index

    price_chf_kwh = aligned * fx_daily.values / 1000.0
    price_chf_kwh.name = "SwissIX[CHF/kWh]"
    return price_chf_kwh


# --------------------------------------------------------------------------
# 2. HT/NT- bzw. SPOT- bzw. Zeitreihen-Tarif
# --------------------------------------------------------------------------

def build_ht_nt_bezugstarif(
    index: pd.DatetimeIndex,
    ht_price: float,
    nt_price: float,
    tage_ht: int,
    start_ht: dt.time,
    end_ht: dt.time,
) -> pd.Series:
    """
    Berechnet den Bezugstarif je Zeitschritt nach HT/NT-Schema.
    tage_ht: Anzahl Tage/Woche mit HT, gezaehlt ab Montag (z.B. 5 -> Mo-Fr HT).
    HT gilt an diesen Tagen im Fenster [start_ht, end_ht), sonst NT.
    """
    is_ht_day = index.weekday < tage_ht
    times = index.time
    is_ht_time = (times >= start_ht) & (times < end_ht)
    is_ht = is_ht_day & is_ht_time
    return pd.Series(
        data=pd.Series(is_ht, index=index).map({True: ht_price, False: nt_price}),
        index=index,
        name="Bezugstarif [CHF/kWh]",
    )


def build_tariffs(
    index: pd.DatetimeIndex,
    params: dict,
    swissix: pd.Series | None,
    existing_ruecklieferung: pd.Series,
    existing_bezug: pd.Series,
) -> tuple[pd.Series, pd.Series]:
    """
    Gibt (rueckliefertarif, bezugstarif) als Series auf `index` zurueck,
    je nach Parameter!C5 (Tarifschema).
    """
    schema = params["tarifschema"]

    if schema == "HT_NT":
        bezug = build_ht_nt_bezugstarif(
            index,
            ht_price=params["ht"],
            nt_price=params["nt"],
            tage_ht=params["tage_ht"],
            start_ht=params["start_ht"],
            end_ht=params["end_ht"],
        )
        ruecklieferung = pd.Series(params["ruecklieferung"], index=index)
        return ruecklieferung, bezug

    elif schema == "SPOT":
        if swissix is None:
            raise ValueError(
                "Tarifschema=SPOT braucht die SwissIX-Zeitreihe (Spalte B) -- "
                "diese ist hier nicht verfuegbar (ENTSO-E-Fetch fehlgeschlagen "
                "oder uebersprungen)."
            )
        bezug = swissix * (1 + params["spot_aufschlag_bezug"] / 100.0)
        ruecklieferung = swissix * (1 - params["spot_abschlag_lieferung"] / 100.0)
        return ruecklieferung, bezug

    elif schema == "Zeitreihen":
        if existing_ruecklieferung.isna().all() or existing_bezug.isna().all():
            warnings.warn(
                "Tarifschema=Zeitreihen erwartet bereits vorhandene historische "
                "Werte in Spalte C/D -- diese sind aber leer. Bitte die "
                "historischen Abnahme-/Bezugstarife von Hand ergaenzen; "
                "das Skript kann diese nicht herleiten."
            )
        return existing_ruecklieferung, existing_bezug

    else:
        raise ValueError(f"Unbekanntes Tarifschema: {schema!r}")


# --------------------------------------------------------------------------
# 3. PV-Referenzprofil (optional, braucht Standortangaben)
# --------------------------------------------------------------------------

def fetch_pv_reference_profile(
    index: pd.DatetimeIndex,
    lat: float,
    lon: float,
    tilt: float = 30,
    azimuth: float = 0,
) -> pd.Series:
    """
    Laedt ein normiertes PV-Erzeugungsprofil (kW pro 1 kWp) via PVGIS
    (EU Joint Research Centre, kostenlos, kein Key noetig) fuer den
    angegebenen Standort. NICHT automatisch aufgerufen -- braucht lat/lon,
    die noch nicht bekannt sind (siehe Kommentar "-> mail Urs/Astrid" im
    Excel). Erst nutzen, wenn Standort/Ausrichtung final feststehen.
    """
    url = "https://re.jrc.ec.europa.eu/api/v5_2/seriescalc"
    params = {
        "lat": lat,
        "lon": lon,
        "startyear": index[0].year,
        "endyear": index[-1].year,
        "pvcalculation": 1,
        "peakpower": 1,  # normiert auf 1 kWp
        "loss": 14,  # Systemverluste in %, Standardannahme
        "angle": tilt,
        "aspect": azimuth,
        "outputformat": "json",
    }
    resp = requests.get(url, params=params, timeout=60)
    resp.raise_for_status()
    hourly = resp.json()["outputs"]["hourly"]
    s = pd.Series(
        {pd.Timestamp(h["time"], format="%Y%m%d:%H%M", utc=True): h["P"] / 1000.0
         for h in hourly}
    ).sort_index()
    s.index = s.index.tz_convert(CET_FIXED)
    return s.reindex(index, method="ffill")


# --------------------------------------------------------------------------
# 4. Parameter einlesen
# --------------------------------------------------------------------------

def read_params(ws_param) -> dict:
    return {
        "beginn": ws_param["C2"].value,
        "ende": ws_param["C3"].value,
        "tarifschema": ws_param["C5"].value,
        "ht": ws_param["C6"].value,
        "nt": ws_param["C7"].value,
        "ruecklieferung": ws_param["C8"].value,
        "tage_ht": int(ws_param["C10"].value),
        "start_ht": ws_param["C11"].value,
        "end_ht": ws_param["C12"].value,
        "spot_abschlag_lieferung": ws_param["C13"].value,
        "spot_aufschlag_bezug": ws_param["C14"].value,
        "dc_leistung": ws_param["C16"].value,
        "produktions_input": ws_param["C17"].value,
    }


# --------------------------------------------------------------------------
# 5. Hauptablauf
# --------------------------------------------------------------------------

def main(
    input_path: str = INPUT_PATH,
    output_path: str = OUTPUT_PATH,
    fetch_swissix: bool = True,
    fetch_pv: bool = False,
    pv_lat: float | None = None,
    pv_lon: float | None = None,
):
    print(f"Lese {input_path} ...")
    wb_values = load_workbook(input_path, data_only=True)   # fuer gecachte Werte (G/H)
    wb = load_workbook(input_path, data_only=False)          # zum Beschreiben/Speichern

    ws_param = wb["Parameter"]
    ws_zeit = wb["Zeitreihen"]
    ws_zeit_vals = wb_values["Zeitreihen"]

    params = read_params(ws_param)
    n_rows = ws_zeit.max_row

    zeit = [ws_zeit_vals.cell(row=r, column=COL_ZEIT).value for r in range(2, n_rows + 1)]
    index = pd.DatetimeIndex(zeit).tz_localize(CET_FIXED)

    # Spalte A war im Original teils eine Array-Formel (Anker in A2, Rest als
    # gespiegelte Werte). openpyxl verliert beim Speichern den gecachten Wert
    # der Formel-Ankerzelle -- deshalb hier explizit mit den eingelesenen,
    # gecachten Zeitstempeln ueberschreiben, damit Spalte A garantiert
    # vollstaendig und formelfrei im Output steht.
    for r, t in zip(range(2, n_rows + 1), zeit):
        ws_zeit.cell(row=r, column=COL_ZEIT).value = t

    existing_ruecklieferung = pd.Series(
        [ws_zeit_vals.cell(row=r, column=COL_RUECKLIEFER).value for r in range(2, n_rows + 1)],
        index=index, dtype=float,
    )
    existing_bezug = pd.Series(
        [ws_zeit_vals.cell(row=r, column=COL_BEZUG).value for r in range(2, n_rows + 1)],
        index=index, dtype=float,
    )

    # --- 1. SwissIX --------------------------------------------------
    swissix = None
    if fetch_swissix:
        try:
            print("Lade SwissIX-Preise von ENTSO-E ...")
            swissix = build_swissix_chf_per_kwh(index)
            for r, v in zip(range(2, n_rows + 1), swissix.values):
                ws_zeit.cell(row=r, column=COL_SWISSIX).value = float(v)
            print("  -> SwissIX erfolgreich ergaenzt.")
        except Exception as exc:  # noqa: BLE001
            warnings.warn(
                f"SwissIX-Fetch fehlgeschlagen ({exc}). Spalte B bleibt leer -- "
                "bitte Skript mit Internetzugang zu web-api.tp.entsoe.eu "
                "erneut ausfuehren."
            )
    else:
        print("SwissIX-Fetch uebersprungen (fetch_swissix=False).")

    # --- 2. Tarife -----------------------------------------------------
    print("Berechne Rueckliefer-/Bezugstarif ...")
    ruecklieferung, bezug = build_tariffs(
        index, params, swissix, existing_ruecklieferung, existing_bezug
    )
    for r, v in zip(range(2, n_rows + 1), ruecklieferung.values):
        if pd.notna(v):
            ws_zeit.cell(row=r, column=COL_RUECKLIEFER).value = float(v)
    for r, v in zip(range(2, n_rows + 1), bezug.values):
        if pd.notna(v):
            ws_zeit.cell(row=r, column=COL_BEZUG).value = float(v)

    # --- 3. PV-Profil (optional) ----------------------------------------
    if fetch_pv and params["produktions_input"] == "normiert":
        if pv_lat is None or pv_lon is None:
            warnings.warn(
                "fetch_pv=True aber pv_lat/pv_lon nicht angegeben -- "
                "PV-Profil (Spalte E) bleibt leer."
            )
        else:
            try:
                print("Lade PV-Referenzprofil von PVGIS ...")
                pv_profile = fetch_pv_reference_profile(index, pv_lat, pv_lon)
                for r, v in zip(range(2, n_rows + 1), pv_profile.values):
                    ws_zeit.cell(row=r, column=COL_PV_NORMIERT).value = float(v)
                print("  -> PV-Profil erfolgreich ergaenzt.")
            except Exception as exc:  # noqa: BLE001
                warnings.warn(f"PVGIS-Fetch fehlgeschlagen ({exc}).")
    else:
        print(
            "PV-Referenzprofil (Spalte E) NICHT ergaenzt -- Standort/Ausrichtung "
            "noch nicht bekannt (siehe Excel-Kommentar 'mail Urs/Astrid'). "
            "main(fetch_pv=True, pv_lat=..., pv_lon=...) aufrufen, sobald bekannt."
        )

    # --- 4. SRL-Preise "backen" (bereits im Original per Formel berechnet) --
    print("Uebernehme SRL-Preise (bereits vorhanden) als feste Werte ...")
    for r in range(2, n_rows + 1):
        for col in (COL_SRL_NEG, COL_SRL_POS):
            cached = ws_zeit_vals.cell(row=r, column=col).value
            if cached is not None:
                ws_zeit.cell(row=r, column=col).value = cached

    # --- 5. Speichern ----------------------------------------------------
    print(f"Speichere {output_path} ...")
    wb.save(output_path)
    print("Fertig.")


if __name__ == "__main__":
    main()